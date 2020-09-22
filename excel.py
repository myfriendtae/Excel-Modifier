
from openpyxl import load_workbook

def make_table(df, writer, filename, sheetname):
    df.to_excel(writer, sheet_name=sheetname, index=False)
    workbook = writer.book
    worksheet = writer.sheets[sheetname]
    (max_row, max_col) = df.shape

    column_settings = []
    for header in df.columns:
        column_settings.append({'header': header})

    worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})
    worksheet.set_column(0, max_col - 1, 12)
    


class Copy_excel:
    def __init__(self,src, destpath):
        self.wb = load_workbook(src)
        self.ws = self.wb["Master"]
        self.dest=destpath

    # Write the value in the cell defined by row_dest+column_dest         
    def write_workbook(self,row_dest,column_dest,value):
        c = self.ws.cell(row = row_dest, column = column_dest)
        c.value = value

    # Save excel file
    def save_excel(self) :  
        self.wb.save(self.dest)